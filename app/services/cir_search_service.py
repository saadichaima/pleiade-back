# app/services/cir_search_service.py
"""
Service de recherche dans les dossiers CIR existants.
- Indexe les documents PDF/Word des équipes
- Lookup du consultant via Excel (avec fuzzy matching)
- Extraction de passages pertinents et articles cités
"""

import os
import re
import hashlib
import pickle
import unicodedata
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field
import openpyxl
from rapidfuzz import fuzz, process

# Pour la lecture de PDF
import fitz  # PyMuPDF

# Pour la lecture de Word
from docx import Document as DocxDocument

# Pour le TF-IDF (fallback)
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np

# Pour les embeddings Azure OpenAI (semantic search)
from Core.embeddings import embed_texts, embed_texts_batch


# Configuration
CIR_BASE_PATH = Path(os.getenv("CIR_FOLDER_PATH", r"C:\Projet bconseil\pleiadeFrontBack\Dossier CIR"))
EXCEL_FILENAME = "JOINTURE BECOME.xlsx"
EMBEDDINGS_CACHE_DIR = CIR_BASE_PATH / ".embeddings_cache"

# Patterns pour détecter les sections "état de l'art" dans les dossiers CIR
ETAT_ART_PATTERNS = [
    r"objet\s+de\s+l[''\u2019]?op[ée]ration\s+de\s+r\s*[&e]\s*d",
    r"description\s+du\s+verrou\s+scientifique",
    r"verrou\s+scientifique\s+ou\s+technique",
    r"[ée]tat\s+de\s+l[''\u2019]?art",
    r"analyse\s+bibliographique",
]


@dataclass
class DocumentChunk:
    """Un passage de texte extrait d'un document."""
    text: str
    document_path: str
    document_name: str
    page_or_section: Optional[str] = None
    team: str = ""
    client_name: str = ""


@dataclass
class CitedArticle:
    """Un article scientifique cité dans un dossier."""
    title: str
    authors: Optional[str] = None
    year: Optional[int] = None
    journal: Optional[str] = None
    source_document: str = ""


@dataclass
class CirSearchResult:
    """Résultat de recherche pour un dossier CIR."""
    client_name: str
    team: str
    consultant: str
    relevant_excerpts: List[Dict] = field(default_factory=list)
    cited_articles: List[CitedArticle] = field(default_factory=list)
    document_paths: List[str] = field(default_factory=list)
    relevance_score: float = 0.0


def normalize_client_name(name: str) -> str:
    """
    Normalise un nom de client pour la comparaison:
    - Supprime accents
    - Mise en minuscules
    - Supprime suffixes juridiques (SAS, SA, SARL, Groupe, etc.)
    - Supprime ponctuation excessive
    """
    if not name:
        return ""

    # Convertir en minuscules
    name = name.lower().strip()

    # Supprimer les accents
    name = unicodedata.normalize('NFKD', name)
    name = ''.join(c for c in name if not unicodedata.combining(c))

    # Supprimer les suffixes juridiques courants
    suffixes = [
        r'\s+sas\s*$', r'\s+sa\s*$', r'\s+sarl\s*$', r'\s+eurl\s*$',
        r'\s+groupe\s*$', r'\s+group\s*$', r'\s+inc\.?\s*$', r'\s+ltd\.?\s*$',
        r'\s+gmbh\s*$', r'\s+s\.?a\.?s\.?\s*$', r'\s+s\.?a\.?\s*$',
        r'\s+france\s*$', r'\s+fr\s*$'
    ]
    for suffix in suffixes:
        name = re.sub(suffix, '', name, flags=re.IGNORECASE)

    # Supprimer ponctuation et espaces multiples
    name = re.sub(r'[^\w\s]', ' ', name)
    name = re.sub(r'\s+', ' ', name).strip()

    return name


class ExcelReferenceLoader:
    """Charge et indexe les données du fichier Excel de référence."""

    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self.client_consultant_map: Dict[str, str] = {}
        self.normalized_map: Dict[str, Tuple[str, str]] = {}  # normalized -> (original_name, consultant)
        self._load_excel()

    def _load_excel(self):
        """Charge le fichier Excel et extrait les mappings client -> consultant."""
        if not self.excel_path.exists():
            print(f"[WARN] Excel file not found: {self.excel_path}")
            return

        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)

            # Parcourir toutes les feuilles
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Trouver les colonnes "client" et "consultant" (ou similaires)
                headers = {}
                for col_idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)), 1):
                    if cell:
                        cell_lower = str(cell).lower().strip()
                        if 'client' in cell_lower or 'societe' in cell_lower or 'société' in cell_lower or 'entreprise' in cell_lower:
                            headers['client'] = col_idx
                        elif 'consultant' in cell_lower or 'charge' in cell_lower or 'responsable' in cell_lower or 'manager' in cell_lower:
                            headers['consultant'] = col_idx

                # Si on n'a pas trouvé les colonnes par nom, essayer les premières colonnes
                if 'client' not in headers:
                    headers['client'] = 1  # Première colonne par défaut
                if 'consultant' not in headers:
                    headers['consultant'] = 2  # Deuxième colonne par défaut

                # Extraire les données
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        client = row[headers['client'] - 1] if headers['client'] - 1 < len(row) else None
                        consultant = row[headers['consultant'] - 1] if headers['consultant'] - 1 < len(row) else None

                        if client and consultant:
                            client_str = str(client).strip()
                            consultant_str = str(consultant).strip()

                            if client_str and consultant_str:
                                self.client_consultant_map[client_str] = consultant_str
                                normalized = normalize_client_name(client_str)
                                self.normalized_map[normalized] = (client_str, consultant_str)
                    except (IndexError, TypeError):
                        continue

            wb.close()
            print(f"[INFO] Loaded {len(self.client_consultant_map)} client-consultant mappings from Excel")

        except Exception as e:
            print(f"[ERROR] Failed to load Excel: {e}")

    def find_consultant(self, client_name: str) -> Tuple[str, str]:
        """
        Trouve le consultant pour un client donné.
        Retourne (consultant_name, match_type) où match_type est 'exact', 'fuzzy' ou 'not_found'.
        """
        if not client_name:
            return ("Non trouvé dans le référentiel", "not_found")

        # 1. Match exact
        if client_name in self.client_consultant_map:
            return (self.client_consultant_map[client_name], "exact")

        # 2. Match exact sur nom normalisé
        normalized = normalize_client_name(client_name)
        if normalized in self.normalized_map:
            return (self.normalized_map[normalized][1], "exact")

        # 3. Fuzzy matching
        if self.normalized_map:
            matches = process.extract(
                normalized,
                list(self.normalized_map.keys()),
                scorer=fuzz.ratio,
                limit=1
            )

            if matches and matches[0][1] >= 75:  # Seuil de 75% de similarité
                matched_normalized = matches[0][0]
                return (self.normalized_map[matched_normalized][1], "fuzzy")

        return ("Non trouvé dans le référentiel", "not_found")


class DocumentIndexer:
    """Indexe et recherche dans les documents CIR."""

    def __init__(self, base_path: Path):
        self.base_path = base_path
        self.documents: List[Dict] = []
        self.chunks: List[DocumentChunk] = []  # uniquement chunks état de l'art
        # TF-IDF (fallback)
        self.vectorizer: Optional[TfidfVectorizer] = None
        self.tfidf_matrix = None
        # Descriptions par document (texte court pour embedding)
        self.doc_descriptions: List[Dict] = []
        # Embeddings sur les descriptions (pas sur les chunks)
        self.embedding_vectors: Optional[np.ndarray] = None
        self.embeddings_available: bool = False
        self._index_documents()

    def _extract_client_from_filename(self, filename: str) -> str:
        """Extrait le nom du client depuis le nom de fichier."""
        # Patterns courants: CLIENT_CIR_24.pdf, CLIENT_MEMOIRE_CIR_24.pdf
        name = Path(filename).stem

        # Supprimer les suffixes courants
        patterns_to_remove = [
            r'_CIR_CII_\d+.*$', r'_CIR_\d+.*$', r'_CII_\d+.*$',
            r'_MEMOIRE_CIR.*$', r'_Memoire_CIR.*$', r'_memoire_cir.*$',
            r'_Mémoire_CIR.*$', r'_Dossier.*$', r'_vF$', r'_VF$', r'_vBC.*$',
            r'_\d{4}$', r'_\d{2}$', r'_\d{2}-\d{2}.*$'
        ]

        for pattern in patterns_to_remove:
            name = re.sub(pattern, '', name, flags=re.IGNORECASE)

        # Remplacer underscores par espaces
        name = name.replace('_', ' ').strip()

        return name

    def _read_pdf(self, path: Path) -> List[Tuple[str, str]]:
        """Lit un PDF et retourne liste de (texte, page_info)."""
        chunks = []
        try:
            doc = fitz.open(str(path))
            for page_num, page in enumerate(doc, 1):
                text = page.get_text()
                if text and len(text.strip()) > 50:
                    chunks.append((text.strip(), f"Page {page_num}"))
            doc.close()
        except Exception as e:
            print(f"[WARN] Error reading PDF {path}: {e}")
        return chunks

    def _read_docx(self, path: Path) -> List[Tuple[str, str]]:
        """Lit un document Word et retourne liste de (texte, section_info)."""
        chunks = []
        try:
            doc = DocxDocument(str(path))
            current_section = "Document"
            current_text = []

            for para in doc.paragraphs:
                # Détecter les titres de section
                if para.style and para.style.name and 'Heading' in para.style.name:
                    # Sauvegarder la section précédente
                    if current_text:
                        text = '\n'.join(current_text)
                        if len(text.strip()) > 50:
                            chunks.append((text.strip(), current_section))
                    current_section = para.text.strip() or "Section"
                    current_text = []
                else:
                    if para.text.strip():
                        current_text.append(para.text.strip())

            # Dernière section
            if current_text:
                text = '\n'.join(current_text)
                if len(text.strip()) > 50:
                    chunks.append((text.strip(), current_section))

        except Exception as e:
            print(f"[WARN] Error reading DOCX {path}: {e}")
        return chunks

    def _extract_cited_articles(self, text: str, source_doc: str) -> List[CitedArticle]:
        """Extrait les références bibliographiques d'un texte."""
        articles = []

        # Patterns pour détecter les citations
        # Pattern ISO: AUTEUR, Prénom. Titre. Journal, année, vol. X, p. Y-Z.
        iso_pattern = r'([A-Z][A-ZÉÈÊËÀÂÄÙÛÜÔÖÎÏÇ]+(?:,?\s+[A-Z][a-zéèêëàâäùûüôöîïç]+\.?)+)[,.]?\s*(.+?)[.,]\s*(\d{4})(?:[,.]|$)'

        # Pattern simple: (Auteur, année) ou [Auteur, année]
        simple_pattern = r'[\[\(]([A-Z][a-zéèêëàâäùûüôöîïç]+(?:\s+et\s+al\.?)?)[,\s]+(\d{4})[\]\)]'

        # Pattern numéroté: [1], [2], etc. avec bibliographie

        seen_titles = set()

        # Recherche ISO
        for match in re.finditer(iso_pattern, text):
            authors = match.group(1).strip()
            title_journal = match.group(2).strip()
            year = int(match.group(3))

            # Séparer titre et journal si possible
            parts = title_journal.split('.')
            title = parts[0].strip() if parts else title_journal
            journal = parts[1].strip() if len(parts) > 1 else None

            if title and title.lower() not in seen_titles and len(title) > 10:
                seen_titles.add(title.lower())
                articles.append(CitedArticle(
                    title=title,
                    authors=authors,
                    year=year,
                    journal=journal,
                    source_document=source_doc
                ))

        # Limiter le nombre d'articles extraits
        return articles[:20]

    def _is_etat_art_content(self, chunk: DocumentChunk) -> bool:
        """Vérifie si un chunk appartient à une section état de l'art / verrou / objet R&D."""
        section_lower = (chunk.page_or_section or "").lower()
        text_start = chunk.text[:500].lower()

        for pattern in ETAT_ART_PATTERNS:
            # Pour DOCX : le titre de section est dans page_or_section
            if re.search(pattern, section_lower):
                return True
            # Pour PDF : le titre peut apparaître dans le texte de la page
            if re.search(pattern, text_start):
                return True

        return False

    def _build_and_embed_descriptions(self):
        """Crée des descriptions courtes par document et les embed."""
        import json

        # Grouper les chunks état de l'art par document
        doc_groups: Dict[str, Dict] = {}
        for i, chunk in enumerate(self.chunks):
            key = chunk.document_path
            if key not in doc_groups:
                doc_groups[key] = {
                    'path': chunk.document_path,
                    'name': chunk.document_name,
                    'client': chunk.client_name,
                    'team': chunk.team,
                    'chunk_indices': [],
                    'text_parts': []
                }
            doc_groups[key]['chunk_indices'].append(i)
            doc_groups[key]['text_parts'].append(chunk.text)

        # Créer une description courte par document (texte tronqué)
        self.doc_descriptions = []
        for doc in doc_groups.values():
            combined = '\n'.join(doc['text_parts'])
            description = combined[:800]  # ~200 tokens, suffisant pour la recherche
            self.doc_descriptions.append({
                'path': doc['path'],
                'client': doc['client'],
                'team': doc['team'],
                'chunk_indices': doc['chunk_indices'],
                'description': description
            })

        # Sauvegarder les descriptions en JSON (inspectable)
        try:
            EMBEDDINGS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
            json_cache = EMBEDDINGS_CACHE_DIR / "descriptions_cache.json"
            json_data = [
                {'client': d['client'], 'team': d['team'], 'description': d['description'][:200]}
                for d in self.doc_descriptions
            ]
            with open(json_cache, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)
            print(f"[CIR] {len(json_data)} descriptions sauvegardées dans {json_cache}")
        except Exception as e:
            print(f"[CIR] Erreur sauvegarde descriptions JSON: {e}")

        # Embed les descriptions (petits batches pour respecter S0 rate limit)
        desc_texts = [d['description'] for d in self.doc_descriptions]
        nb_docs = len(desc_texts)
        desc_hash = self._compute_chunks_hash(desc_texts)

        cached_vectors = self._load_embedding_cache(desc_hash, nb_docs)
        if cached_vectors is not None:
            self.embedding_vectors = cached_vectors
            self.embeddings_available = True
        else:
            print(f"[CIR] Embedding de {nb_docs} descriptions (~800 chars chacune)...")
            vectors = embed_texts_batch(desc_texts, batch_size=20, delay_between=2)
            if vectors:
                self.embedding_vectors = np.array(vectors, dtype=np.float32)
                self.embeddings_available = True
                self._save_embedding_cache(self.embedding_vectors, desc_hash, nb_docs)
            else:
                print("[CIR] Embeddings indisponibles, fallback sur TF-IDF")
                self.embeddings_available = False

    def _compute_chunks_hash(self, texts: List[str]) -> str:
        """Calcule un hash SHA256 des textes pour détecter les changements."""
        h = hashlib.sha256()
        for t in texts:
            h.update(t.encode('utf-8', errors='ignore'))
        return h.hexdigest()

    def _load_embedding_cache(self, chunks_hash: str, expected_count: int) -> Optional[np.ndarray]:
        """Charge les embeddings depuis le cache disque si le hash correspond."""
        cache_path = EMBEDDINGS_CACHE_DIR / "cache.pkl"
        if not cache_path.exists():
            return None
        try:
            with open(cache_path, 'rb') as f:
                cache = pickle.load(f)
            if cache.get("chunks_hash") == chunks_hash and cache.get("doc_count") == expected_count:
                print(f"[CIR] Cache embeddings valide, chargement de {cache['doc_count']} vecteurs")
                return cache["vectors"]
            else:
                print("[CIR] Cache embeddings invalide (documents modifiés), recalcul nécessaire")
                return None
        except Exception as e:
            print(f"[CIR] Erreur lecture cache: {e}")
            return None

    def _save_embedding_cache(self, vectors: np.ndarray, chunks_hash: str, doc_count: int):
        """Sauvegarde les embeddings en cache disque."""
        try:
            EMBEDDINGS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
            cache_path = EMBEDDINGS_CACHE_DIR / "cache.pkl"
            cache = {
                "chunks_hash": chunks_hash,
                "vectors": vectors,
                "doc_count": doc_count
            }
            with open(cache_path, 'wb') as f:
                pickle.dump(cache, f)
            print(f"[CIR] Embeddings sauvegardés en cache ({doc_count} vecteurs)")
        except Exception as e:
            print(f"[CIR] Erreur sauvegarde cache: {e}")

    def _index_documents(self):
        """Indexe les documents CIR en filtrant uniquement les sections état de l'art."""
        if not self.base_path.exists():
            print(f"[WARN] CIR base path not found: {self.base_path}")
            return

        # Phase 1 : Lire tous les documents
        all_chunks: List[DocumentChunk] = []

        for team_dir in self.base_path.iterdir():
            if not team_dir.is_dir() or not team_dir.name.startswith("Equipe"):
                continue

            team_name = team_dir.name

            for file_path in team_dir.iterdir():
                if file_path.suffix.lower() in ['.pdf', '.docx', '.doc']:
                    client_name = self._extract_client_from_filename(file_path.name)

                    if file_path.suffix.lower() == '.pdf':
                        raw_chunks = self._read_pdf(file_path)
                    else:
                        raw_chunks = self._read_docx(file_path)

                    for text, location in raw_chunks:
                        all_chunks.append(DocumentChunk(
                            text=text,
                            document_path=str(file_path),
                            document_name=file_path.name,
                            page_or_section=location,
                            team=team_name,
                            client_name=client_name
                        ))

                    self.documents.append({
                        'path': str(file_path),
                        'name': file_path.name,
                        'team': team_name,
                        'client': client_name
                    })

        # Phase 2 : Filtrer — garder uniquement les sections état de l'art
        etat_art_chunks = [c for c in all_chunks if self._is_etat_art_content(c)]

        if etat_art_chunks:
            self.chunks = etat_art_chunks
            print(f"[CIR] {len(self.chunks)} chunks état de l'art trouvés (sur {len(all_chunks)} total)")
        else:
            self.chunks = all_chunks
            print(f"[CIR] Aucun chunk état de l'art détecté, utilisation de tous les {len(all_chunks)} chunks")

        if not self.chunks:
            print("[WARN] No chunks found")
            return

        # Phase 3 : TF-IDF sur les chunks filtrés (fallback)
        texts = [chunk.text for chunk in self.chunks]
        self.vectorizer = TfidfVectorizer(
            max_features=10000,
            ngram_range=(1, 2),
            stop_words=None
        )
        self.tfidf_matrix = self.vectorizer.fit_transform(texts)

        # Phase 4 : Descriptions courtes par document + embeddings
        self._build_and_embed_descriptions()

        mode = "semantic (descriptions)" if self.embeddings_available else "TF-IDF (fallback)"
        print(f"[INFO] Indexed {len(self.documents)} documents, {len(self.chunks)} chunks état de l'art, {len(self.doc_descriptions)} descriptions — mode: {mode}")

    def search(self, query: str, top_k: int = 10) -> List[Tuple[DocumentChunk, float]]:
        """Recherche principale : semantic si dispo, sinon TF-IDF."""
        if self.embeddings_available:
            return self.search_semantic(query, top_k)
        return self.search_tfidf(query, top_k)

    def search_semantic(self, query: str, top_k: int = 10) -> List[Tuple[DocumentChunk, float]]:
        """Recherche sémantique via embeddings sur les descriptions de documents."""
        if self.embedding_vectors is None or not self.doc_descriptions:
            return self.search_tfidf(query, top_k)

        # Embed la requête
        query_vecs = embed_texts([query])
        if not query_vecs:
            print("[CIR] Échec embedding requête, fallback TF-IDF")
            return self.search_tfidf(query, top_k)

        query_vec = query_vecs[0].reshape(1, -1)

        # Cosine similarity sur les descriptions
        similarities = cosine_similarity(query_vec, self.embedding_vectors).flatten()
        top_doc_indices = similarities.argsort()[::-1]

        # Retourner les chunks état de l'art des documents matchés
        results = []
        for doc_idx in top_doc_indices:
            doc_score = float(similarities[doc_idx])
            if doc_score < 0.3:
                break
            for chunk_idx in self.doc_descriptions[doc_idx]['chunk_indices']:
                results.append((self.chunks[chunk_idx], doc_score))
            if len(results) >= top_k:
                break

        return results[:top_k]

    def search_tfidf(self, query: str, top_k: int = 10) -> List[Tuple[DocumentChunk, float]]:
        """Recherche TF-IDF (fallback)."""
        if not self.vectorizer or self.tfidf_matrix is None:
            return []

        query_vec = self.vectorizer.transform([query])
        similarities = cosine_similarity(query_vec, self.tfidf_matrix).flatten()
        top_indices = similarities.argsort()[-top_k:][::-1]

        results = []
        for idx in top_indices:
            if similarities[idx] > 0.05:
                results.append((self.chunks[idx], float(similarities[idx])))

        return results


class CirSearchService:
    """Service principal de recherche CIR."""

    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if self._initialized:
            return

        self.base_path = CIR_BASE_PATH
        self.excel_loader = ExcelReferenceLoader(self.base_path / EXCEL_FILENAME)
        self.doc_indexer = DocumentIndexer(self.base_path)
        self._initialized = True

    def search(self, query: str, max_results: int = 5) -> List[CirSearchResult]:
        """
        Recherche des dossiers CIR similaires à la requête.

        Args:
            query: Description du projet (problématique, verrous, techno, contexte)
            max_results: Nombre max de résultats à retourner (3-5)

        Returns:
            Liste de CirSearchResult avec projets comparables
        """
        # Recherche dans les documents
        search_results = self.doc_indexer.search(query, top_k=50)

        # Grouper par client/projet
        results_by_client: Dict[str, CirSearchResult] = {}

        for chunk, score in search_results:
            client_key = f"{chunk.client_name}_{chunk.team}"

            if client_key not in results_by_client:
                # Lookup consultant
                consultant, match_type = self.excel_loader.find_consultant(chunk.client_name)

                results_by_client[client_key] = CirSearchResult(
                    client_name=chunk.client_name,
                    team=chunk.team,
                    consultant=consultant,
                    relevant_excerpts=[],
                    cited_articles=[],
                    document_paths=[],
                    relevance_score=0.0
                )

            result = results_by_client[client_key]

            # Ajouter l'extrait
            if len(result.relevant_excerpts) < 3:  # Max 3 extraits par projet
                # Tronquer le texte si trop long
                excerpt_text = chunk.text[:1500] + "..." if len(chunk.text) > 1500 else chunk.text

                result.relevant_excerpts.append({
                    'text': excerpt_text,
                    'document': chunk.document_name,
                    'location': chunk.page_or_section,
                    'score': score
                })

            # Mettre à jour le score (moyenne pondérée)
            result.relevance_score = max(result.relevance_score, score)

            # Ajouter le document path
            if chunk.document_path not in result.document_paths:
                result.document_paths.append(chunk.document_path)

            # Extraire les articles cités
            if len(result.cited_articles) < 10:
                articles = self.doc_indexer._extract_cited_articles(chunk.text, chunk.document_name)
                for article in articles:
                    if not any(a.title.lower() == article.title.lower() for a in result.cited_articles):
                        result.cited_articles.append(article)

        # Trier par score et limiter
        sorted_results = sorted(
            results_by_client.values(),
            key=lambda r: r.relevance_score,
            reverse=True
        )[:max_results]

        return sorted_results

    def get_stats(self) -> Dict:
        """Retourne des statistiques sur l'index."""
        return {
            'total_documents': len(self.doc_indexer.documents),
            'total_chunks': len(self.doc_indexer.chunks),
            'total_clients_in_excel': len(self.excel_loader.client_consultant_map),
            'teams': list(set(d['team'] for d in self.doc_indexer.documents))
        }


# Singleton pour réutilisation
def get_cir_search_service() -> CirSearchService:
    return CirSearchService()
